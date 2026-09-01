import pandas as pd
from openpyxl import load_workbook

# Master keys
master_150 = pd.read_csv("review_150_master_key_INTERNAL.csv")
master_D = pd.read_csv("review_batchD_master_key_INTERNAL.csv")

batch_files = {
    "BatchA_IndicConformer": "BatchA_IndicConformer_evaluated.xlsx",
    "BatchB_Whisper": "BatchB_Whisper_evaluated.xlsx",
    "BatchC_Qwen": "BatchC_Qwen_evaluated.xlsx",
    "BatchD_QwenNormalized": "BatchD_QwenNormalized_evaluated.xlsx",
}

HUMAN_TO_AUTO = {
    "1 - Correct / Very Close": "FAITHFUL",
    "2 - Mistake but Related": "PARAPHRASE_OR_RECOGNITION_ERROR",
    "3 - Made up / Unrelated": "HALLUCINATION_CANDIDATE",
}

def extract_human_labels(path):
    wb = load_workbook(path, data_only=True)
    ws = wb["Annotation"]
    rows = []
    for r in range(3, ws.max_row + 1):
        row_id = ws.cell(row=r, column=1).value
        label = ws.cell(row=r, column=5).value
        notes = ws.cell(row=r, column=6).value
        prediction = ws.cell(row=r, column=4).value
        if row_id is None:
            continue
        # Treat "empty output" noted (or blank prediction with no dropdown match) as its own valid label
        is_blank_label = label is None or str(label).strip() == ""
        note_says_empty = notes is not None and "empty" in str(notes).lower()
        pred_is_empty = prediction is None or str(prediction).strip() == ""
        if is_blank_label and (note_says_empty or pred_is_empty):
            label = "EMPTY_OUTPUT_HUMAN"
        rows.append({"review_row_id": int(row_id), "human_label_raw": label, "notes": notes})
    return pd.DataFrame(rows)

all_merged = []
for batch, fname in batch_files.items():
    human_df = extract_human_labels(fname)
    key = master_D if batch == "BatchD_QwenNormalized" else master_150[master_150["batch"] == batch]
    merged = key.merge(human_df, on="review_row_id", how="left")
    merged["batch"] = batch
    all_merged.append(merged)

full = pd.concat(all_merged, ignore_index=True)

def map_human_label(raw):
    if raw == "EMPTY_OUTPUT_HUMAN":
        return "EMPTY_OUTPUT"
    return HUMAN_TO_AUTO.get(raw)

full["human_label_mapped"] = full["human_label_raw"].apply(map_human_label)
full["has_human_label"] = full["human_label_mapped"].notna()
full["agrees"] = full["human_label_mapped"] == full["label_semantic"]

full.to_csv("human_validation_merged.csv", index=False)

print(f"Total rows: {len(full)}")
print(f"Rows with a human label: {full['has_human_label'].sum()}")
print(f"Rows left blank by annotator: {(~full['has_human_label']).sum()}")
print()

scored = full[full["has_human_label"]]
overall_agreement = scored["agrees"].mean()
print(f"OVERALL agreement (human vs automatic classifier): {overall_agreement:.3f}  (n={len(scored)})")
print()

print("Agreement by batch / model:")
for batch in scored["batch"].unique():
    sub = scored[scored["batch"] == batch]
    print(f"  {batch}: {sub['agrees'].mean():.3f}  (n={len(sub)})")

print()
print("Confusion: automatic label vs human label (counts)")
confusion = pd.crosstab(scored["label_semantic"], scored["human_label_mapped"])
print(confusion)

print()
print("Notes left by annotators (non-empty):")
notes = scored[scored["notes"].notna() & (scored["notes"].astype(str).str.strip() != "")]
for _, row in notes.iterrows():
    print(f"  [{row['batch']} row {row['review_row_id']}] auto={row['label_semantic']} human={row['human_label_mapped']} :: {row['notes']}")
